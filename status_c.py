# -*- coding: utf-8 -*-

import sys
import time
import datetime
import openpyxl
# from openpyxl import Workbook
import csv
from lib import IN_NAME, IN_SNILS, IN_STAT_OUR, IN_STAT_FOND , OUT_NAME, OUT_STAT, OUT_FOND_PAY, lenl

workbooks =  []
sheets = []
for i, xlsx_file in enumerate(sys.argv):                              # Загружаем все xlsx файлы
    if i == 0:
        continue
    workbooks.append(openpyxl.load_workbook(filename=xlsx_file, read_only=True))
    sheets.append(workbooks[i-1][workbooks[i-1].sheetnames[0]])
#    for j, row in enumerate(sheets[i-1].rows):
#        for k, cell in enumerate(row):
#            g=0

sheets_keys = []
for i, sheet in enumerate(sheets):                                    # Маркируем нужные столбцы
    keys = {}
    for j, row in enumerate(sheet.rows):
        if j > 0:
            break
        for k, cell in enumerate(row):                                # Проверяем, чтобы был СНИЛС
            if cell.value in IN_SNILS:
                keys[IN_SNILS[0]] = k
        if len(keys) > 0:
            for k, cell in enumerate(row):
                for n, name in enumerate(IN_NAME):
                    if n == 0:
                        continue
                    if cell.value != None:
                        if cell.value == name:
                            keys[name] = k
        else:
            print('В файле ' + sys.argv[i+1] + 'отсутствует колонка со СНИЛС')
            time.sleep(3)
            sys.exit()
    sheets_keys.append(keys)

without = True
for i, sheet in enumerate(sheets):
    if len(sheets_keys[i]) > 1:
        print('\nВ файле ' + sys.argv[i+1] + ' найдены столбцы:')
        for q in sheets_keys[i].keys():
            print('    ' + q)
        without = False
if without:
    print('Во всех файлах нет никаких столбцов, кроме СНИЛС')
    time.sleep(3)
    sys.exit()

print('\n'+ datetime.datetime.now().strftime("%H:%M:%S") +'Начинаем расчет \n')

our_statuses = []
fond_pays = []
total_rows = sheets[0].max_row
perc_rows = 0
for j, row in enumerate(sheets[0].rows):                     # Загружаем все входные данные в одну строку
    our_status = {}
    fond_pay = {}
    if j == 0:
        continue
    big_row = {}
    if lenl(row[sheets_keys[0][IN_SNILS[0]]].value) != 11:
        continue
    for k, sheet_key in enumerate(sheets_keys[0]):                              # Из первого файла
        if row[sheets_keys[0][sheet_key]].value != None \
                and str(row[sheets_keys[0][sheet_key]].value).strip() != '':
              # and str(row[sheets_keys[0][sheet_key]].value).strip() != '—'
            big_row[sheet_key] = str(row[sheets_keys[0][sheet_key]].value)
    for i, sheet in enumerate(sheets):                                          # Из всех остальных
        if i == 0:
            continue
        if str(type(big_row[IN_SNILS[0]])) == "<class 'str'>":
            if big_row[IN_SNILS[0]].strip() != '':
                for row in sheets[i].rows:
                    if str(type(row[sheets_keys[i][IN_SNILS[0]]].value)) == "<class 'str'>":
                        if row[sheets_keys[i][IN_SNILS[0]]].value.strip() == big_row[IN_SNILS[0]].strip():
                            for k, sheet_key in enumerate(sheets_keys[i]):
                                tek = row[sheets_keys[i][sheet_key]].value
                                if tek != None and str(tek).strip() != '':                # and str(tek).strip() != '—'
                                    if str(tek)[:12].lower() == 'одобрено инф':
                                        big_row[sheet_key] = OUT_STAT['Фонд - Статус КоллЦентра'][13]
                                    elif str(tek)[:15].lower() == 'акцепт прозвона':
                                        big_row[sheet_key] = OUT_STAT['Фонд - Статус КоллЦентра'][12]
                                    else:
                                        big_row[sheet_key] = str(row[sheets_keys[i][sheet_key]].value)
                            break


    for i, name in enumerate(OUT_NAME):                                 # Заполняем our_status пустыми значениями(None)
        try:
            if our_status[name] == None:
                q = 0
        except KeyError:
            our_status[name] = None

    for i, name in enumerate(OUT_FOND_PAY):                                 # Заполняем fond_pays пустыми значениями(None)
        try:
            if fond_pay[name] == None:
                q = 0
        except KeyError:
            fond_pay[name] = None


    for i, name in enumerate(IN_NAME):          # Заполняем строку списка-словаря для csv файла статусами фонда
        if i == 0 :
            our_status[name] = big_row[name]    # СНИЛС
            fond_pay[name] = big_row[name]

# Вручную, Бумага принята только если в обоих полях Исправили или Наличие бумаги
        elif name == 'СтатусБумажногоНосителяПоДоговору' or name == 'СтатусБумажногоНосителяПоЗаявлению':
            try:
                tek1 = big_row['СтатусБумажногоНосителяПоДоговору']
                tek2 = big_row['СтатусБумажногоНосителяПоЗаявлению']
                if (tek1 == 'Исправили' or tek1 == 'Наличие бумаги') \
                                and (tek2 == 'Исправили' or tek2 == 'Наличие бумаги'):
                    our_status['Фонд - Статус бумаги'] = OUT_STAT['Фонд - Статус бумаги'].index('Бумага принята')
                else:
                    our_status['Фонд - Статус бумаги'] = OUT_STAT['Фонд - Статус бумаги'].index('Ошибка')
            except KeyError:
                our_status['Фонд - Статус бумаги'] = OUT_STAT['Фонд - Статус бумаги'].index('Ошибка')
            except ValueError:
                q= 0
        elif name == 'СтатусОплаты':
            try:
                fond_pay[IN_STAT_FOND[name][big_row[name]][0]] = \
                    OUT_STAT[IN_STAT_FOND[name][big_row[name]][0]].index(IN_STAT_FOND[name][big_row[name]][1])
            except KeyError:
                q = 0
            except ValueError:
                q = 0
        else:
            try:
                our_status[IN_STAT_FOND[name][big_row[name]][0]] = \
                    OUT_STAT[IN_STAT_FOND[name][big_row[name]][0]].index(IN_STAT_FOND[name][big_row[name]][1])
            except KeyError:
                q = 0
            except ValueError:
                q= 0
# Заполняем строку списка-словаря для csv файла нашими статусами, если нет из фонда
    for i, name in enumerate(IN_NAME):
        if i == 0:
            continue
        else:
            try:
                if our_status[IN_STAT_OUR[name][big_row[name]][0]] == None:
                    our_status[IN_STAT_OUR[name][big_row[name]][0]] = \
                        OUT_STAT[IN_STAT_OUR[name][big_row[name]][0]].index(IN_STAT_OUR[name][big_row[name]][1])
            except KeyError:
                q = 0
            except ValueError:
                q = 0

    without = True
    for i, stat in enumerate(our_status):
        if i == 0:
            continue
        if stat != None:
            without = False
    p_status = 'Пропущены'
    p_pay = 'Не оплачено - строка пропущена'
    if not without:
        our_statuses.append(our_status)
        p_status = 'Добавлены'

    if fond_pay[OUT_FOND_PAY[1]] != None:
        fond_pays.append(fond_pay)
        p_pay = 'Оплачено - строка добавлена'

    if int(j/total_rows*100) > perc_rows:
        perc_rows = int(j/total_rows*100)
        print(datetime.datetime.now().strftime("%H:%M:%S") + '  обработано ' + str(perc_rows) + '%')

#    print(big_row['СНИЛС'] + ' Статусы: ' + p_status + ' Оплата: ' + p_pay)

# our_statuses = [{'Имя':'Он они он','Возраст':25,'Вес':200}, {'Имя':'Я я я','Возраст':31,'Вес':180}]
with open('statuses.csv', 'w', encoding='cp1251') as output_file:
    dict_writer = csv.DictWriter(output_file, OUT_NAME, delimiter=';') #, quoting=csv.QUOTE_NONNUMERIC)
    dict_writer.writeheader()
    dict_writer.writerows(our_statuses)
output_file.close()
with open('fond_pays.csv', 'w', encoding='cp1251') as output_file:
    dict_writer = csv.DictWriter(output_file, OUT_FOND_PAY, delimiter=';') #, quoting=csv.QUOTE_NONNUMERIC)
    dict_writer.writeheader()
    dict_writer.writerows(fond_pays)
output_file.close()
